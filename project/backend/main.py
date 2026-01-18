from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import requests
import os
from openpyxl import Workbook, load_workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

FILE = "users.xlsx"


class User(BaseModel):
    name: str
    phone: str
    zone: str


class RouteRequest(BaseModel):
    start: list  
    end: list    


@app.post("/register")
def register_user(user: User):
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Zone"])
        wb.save(FILE)

    wb = load_workbook(FILE)
    ws = wb.active
    ws.append([user.name, user.phone, user.zone])
    wb.save(FILE)

    return {"status": "saved"}


@app.post("/route")
def route(data: RouteRequest):
    try:
        s = f"{data.start[1]},{data.start[0]}"
        e = f"{data.end[1]},{data.end[0]}"
        url = (
            f"http://localhost:5000/route/v1/driving/"
            f"{s};{e}?overview=full&geometries=geojson"
        )

        r = requests.get(url, timeout=5).json()
        route_data = r["routes"][0]
        coords = [[c[1], c[0]] for c in route_data["geometry"]["coordinates"]]

        return {
            "route": coords,
            "distance": round(route_data["distance"] / 1000, 2),
            "duration": round(route_data["duration"] / 60, 2)
        }
    except Exception:
        return {"error": "OSRM not reachable"}


@app.get("/")
def root():
    return {"status": "API running"}


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
